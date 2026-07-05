from flask import Flask, jsonify, request, render_template, redirect, url_for, send_file
from config import Config
from models import db, Product, Sale, PaymentMethod, Paybill, ProductVariant
from flask_migrate import Migrate
from sqlalchemy import func
from datetime import datetime, date
import io

app = Flask(__name__)
app.config.from_object(Config)
db.init_app(app)
migrate = Migrate(app, db)

# ─────────────────────────────────────────
# HEALTH
# ─────────────────────────────────────────
@app.route('/api/health')
def health():
    return jsonify({'status': 'healthy'})

# ─────────────────────────────────────────
# CATEGORIES
# ─────────────────────────────────────────
@app.route('/api/categories', methods=['GET'])
def get_categories():
    rows = db.session.query(Product.category).filter(Product.is_active == True).distinct().order_by(Product.category).all()
    return jsonify([r[0] for r in rows])

# ─────────────────────────────────────────
# PRODUCTS
# ─────────────────────────────────────────
@app.route('/api/products', methods=['GET', 'POST'])
def products():
    if request.method == 'POST':
        data = request.get_json()
        new_product = Product(
            product_name=data['product_name'],
            description=data.get('description'),
            category=data['category'],
            selling_price=data['selling_price'],
            buying_price=data['buying_price'],
            quantity_in_stock=data.get('quantity_in_stock', 0),
            low_stock_threshold=data.get('low_stock_threshold', 5),
            image_path=data.get('image_path')
        )
        db.session.add(new_product)
        db.session.commit()
        return jsonify(new_product.to_dict()), 201

    show_discontinued = request.args.get('show_discontinued', 'false').lower() == 'true'
    category_filter = request.args.get('category', '').strip()

    query = Product.query
    if not show_discontinued:
        query = query.filter(Product.is_active == True)
    if category_filter:
        query = query.filter(Product.category.ilike(f'%{category_filter}%'))
    return jsonify([p.to_dict() for p in query.all()])


@app.route('/api/products/<int:product_id>', methods=['GET', 'PUT', 'DELETE'])
def product_detail(product_id):
    product = Product.query.get_or_404(product_id)

    if request.method == 'GET':
        return jsonify(product.to_dict())

    elif request.method == 'PUT':
        data = request.get_json()
        product.product_name = data.get('product_name', product.product_name)
        product.description = data.get('description', product.description)
        product.category = data.get('category', product.category)
        product.selling_price = data.get('selling_price', product.selling_price)
        product.buying_price = data.get('buying_price', product.buying_price)
        product.quantity_in_stock = data.get('quantity_in_stock', product.quantity_in_stock)
        product.low_stock_threshold = data.get('low_stock_threshold', product.low_stock_threshold)
        db.session.commit()
        return jsonify(product.to_dict())

    elif request.method == 'DELETE':
        if product.sales:
            return jsonify({
                'error': f'Cannot delete "{product.product_name}" — it has sales records attached.'
            }), 409
        db.session.delete(product)
        db.session.commit()
        return jsonify({'message': 'Product deleted'}), 200


# ─────────────────────────────────────────
# DISCONTINUE / RESTORE
# ─────────────────────────────────────────
@app.route('/api/products/<int:product_id>/discontinue', methods=['PUT'])
def discontinue_product(product_id):
    product = Product.query.get_or_404(product_id)
    product.is_active = False
    db.session.commit()
    return jsonify({'message': f'"{product.product_name}" has been discontinued.', 'product': product.to_dict()})


@app.route('/api/products/<int:product_id>/restore', methods=['PUT'])
def restore_product(product_id):
    product = Product.query.get_or_404(product_id)
    product.is_active = True
    db.session.commit()
    return jsonify({'message': f'"{product.product_name}" has been restored.', 'product': product.to_dict()})


# ─────────────────────────────────────────
# PRODUCT VARIANTS
# ─────────────────────────────────────────
@app.route('/api/products/<int:product_id>/variants', methods=['GET', 'POST'])
def product_variants(product_id):
    product = Product.query.get_or_404(product_id)

    if request.method == 'POST':
        data = request.get_json()
        variant = ProductVariant(
            product_id=product_id,
            size=data.get('size'),
            color=data.get('color'),
            material=data.get('material'),
            selling_price=data['selling_price'],
            buying_price=data['buying_price'],
            quantity_in_stock=data.get('quantity_in_stock', 0),
            low_stock_threshold=data.get('low_stock_threshold', 5)
        )
        db.session.add(variant)
        db.session.commit()
        return jsonify(variant.to_dict()), 201

    variants = ProductVariant.query.filter_by(product_id=product_id, is_active=True).all()
    return jsonify([v.to_dict() for v in variants])


@app.route('/api/variants/<int:variant_id>', methods=['GET', 'PUT'])
def variant_detail(variant_id):
    variant = ProductVariant.query.get_or_404(variant_id)

    if request.method == 'GET':
        return jsonify(variant.to_dict())

    data = request.get_json()
    variant.size                = data.get('size', variant.size)
    variant.color               = data.get('color', variant.color)
    variant.material            = data.get('material', variant.material)
    variant.selling_price       = data.get('selling_price', variant.selling_price)
    variant.buying_price        = data.get('buying_price', variant.buying_price)
    variant.quantity_in_stock   = data.get('quantity_in_stock', variant.quantity_in_stock)
    variant.low_stock_threshold = data.get('low_stock_threshold', variant.low_stock_threshold)
    db.session.commit()
    return jsonify(variant.to_dict())


@app.route('/api/variants/<int:variant_id>/discontinue', methods=['PUT'])
def discontinue_variant(variant_id):
    variant = ProductVariant.query.get_or_404(variant_id)
    variant.is_active = False
    db.session.commit()
    return jsonify({'message': 'Variant discontinued.', 'variant': variant.to_dict()})


@app.route('/api/variants/<int:variant_id>/restore', methods=['PUT'])
def restore_variant(variant_id):
    variant = ProductVariant.query.get_or_404(variant_id)
    variant.is_active = True
    db.session.commit()
    return jsonify({'message': 'Variant restored.', 'variant': variant.to_dict()})


# ─────────────────────────────────────────
# SALES
# ─────────────────────────────────────────
@app.route('/api/sales', methods=['GET', 'POST'])
def sales():
    if request.method == 'POST':
        data = request.get_json()
        product = Product.query.get_or_404(data['product_id'])

        quantity   = data['quantity_sold']
        variant_id = data.get('variant_id')
        variant    = None

        if variant_id:
            variant = ProductVariant.query.get_or_404(variant_id)
            if variant.product_id != product.product_id:
                return jsonify({'error': 'Variant does not belong to this product.'}), 400
            if variant.quantity_in_stock < quantity:
                return jsonify({'error': 'Insufficient variant stock.'}), 400
            selling_price = float(variant.selling_price)
            buying_price  = float(variant.buying_price)
        else:
            if product.quantity_in_stock < quantity:
                return jsonify({'error': 'Insufficient stock.'}), 400
            selling_price = float(product.selling_price)
            buying_price  = float(product.buying_price)

        total_revenue = quantity * selling_price
        total_cost    = quantity * buying_price
        profit        = total_revenue - total_cost

        new_sale = Sale(
            product_id=data['product_id'],
            variant_id=variant_id,
            quantity_sold=quantity,
            selling_price_at_sale=selling_price,
            buying_price_at_sale=buying_price,
            total_revenue=total_revenue,
            total_cost=total_cost,
            profit=profit,
            payment_method_id=data['payment_method_id'],
            paybill_id=data.get('paybill_id')
        )

        if variant:
            variant.quantity_in_stock -= quantity
        else:
            product.quantity_in_stock -= quantity

        db.session.add(new_sale)
        db.session.commit()
        return jsonify(new_sale.to_dict()), 201

    sales_list = Sale.query.all()
    return jsonify([s.to_dict() for s in sales_list])


# ─────────────────────────────────────────
# REPORTS — overall summary
# ─────────────────────────────────────────
@app.route('/api/reports')
def reports():
    totals = db.session.query(
        func.sum(Sale.total_revenue).label('total_revenue'),
        func.sum(Sale.total_cost).label('total_cost'),
        func.sum(Sale.profit).label('total_profit'),
        func.sum(Sale.quantity_sold).label('total_items_sold')
    ).first()

    low_stock = Product.query.filter(
        Product.quantity_in_stock <= Product.low_stock_threshold
    ).all()

    category_sales = db.session.query(
        Product.category,
        func.sum(Sale.total_revenue).label('revenue'),
        func.sum(Sale.profit).label('profit')
    ).join(Sale).group_by(Product.category).all()

    return jsonify({
        'summary': {
            'total_revenue': float(totals.total_revenue or 0),
            'total_cost': float(totals.total_cost or 0),
            'total_profit': float(totals.total_profit or 0),
            'total_items_sold': int(totals.total_items_sold or 0)
        },
        'low_stock_products': [p.to_dict() for p in low_stock],
        'sales_by_category': [
            {'category': c, 'revenue': float(r), 'profit': float(p)}
            for c, r, p in category_sales
        ]
    })


# ─────────────────────────────────────────
# REPORTS — date range
# ─────────────────────────────────────────
@app.route('/api/reports/range', methods=['GET'])
def range_report():
    start_str = request.args.get('start')
    end_str = request.args.get('end')

    if not start_str or not end_str:
        return jsonify({'error': 'Both start and end dates are required (YYYY-MM-DD).'}), 400

    try:
        start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD.'}), 400

    if start_date > end_date:
        return jsonify({'error': 'Start date must be before end date.'}), 400

    earliest = db.session.query(func.min(func.date(Sale.sale_date))).scalar()
    if earliest and start_date < earliest:
        return jsonify({
            'error': f'Start date is before the earliest sale record in the system ({earliest}).'
        }), 400

    totals = db.session.query(
        func.sum(Sale.total_revenue).label('total_revenue'),
        func.sum(Sale.total_cost).label('total_cost'),
        func.sum(Sale.profit).label('total_profit'),
        func.sum(Sale.quantity_sold).label('total_items_sold'),
        func.count(Sale.sale_id).label('total_transactions')
    ).filter(
        func.date(Sale.sale_date) >= start_date,
        func.date(Sale.sale_date) <= end_date
    ).first()

    sales_in_range = Sale.query.filter(
        func.date(Sale.sale_date) >= start_date,
        func.date(Sale.sale_date) <= end_date
    ).order_by(Sale.sale_date.asc()).all()

    category_breakdown = db.session.query(
        Product.category,
        func.sum(Sale.total_revenue).label('revenue'),
        func.sum(Sale.profit).label('profit'),
        func.sum(Sale.quantity_sold).label('items_sold')
    ).join(Sale, Product.product_id == Sale.product_id).filter(
        func.date(Sale.sale_date) >= start_date,
        func.date(Sale.sale_date) <= end_date
    ).group_by(Product.category).all()

    payment_breakdown = db.session.query(
        PaymentMethod.method_name,
        func.sum(Sale.total_revenue).label('revenue'),
        func.count(Sale.sale_id).label('transactions')
    ).join(Sale, PaymentMethod.payment_method_id == Sale.payment_method_id).filter(
        func.date(Sale.sale_date) >= start_date,
        func.date(Sale.sale_date) <= end_date
    ).group_by(PaymentMethod.method_name).all()

    return jsonify({
        'start_date': start_str,
        'end_date': end_str,
        'summary': {
            'total_revenue': float(totals.total_revenue or 0),
            'total_cost': float(totals.total_cost or 0),
            'total_profit': float(totals.total_profit or 0),
            'total_items_sold': int(totals.total_items_sold or 0),
            'total_transactions': int(totals.total_transactions or 0)
        },
        'sales': [s.to_dict() for s in sales_in_range],
        'category_breakdown': [
            {'category': c, 'revenue': float(r), 'profit': float(p), 'items_sold': int(i)}
            for c, r, p, i in category_breakdown
        ],
        'payment_breakdown': [
            {'method': m, 'revenue': float(r), 'transactions': int(t)}
            for m, r, t in payment_breakdown
        ]
    })


# ─────────────────────────────────────────
# REPORTS — export PDF or Excel
# ─────────────────────────────────────────
@app.route('/api/reports/export', methods=['GET'])
def export_report():
    fmt = request.args.get('format', 'pdf').lower()
    start_str = request.args.get('start')
    end_str = request.args.get('end')

    query = Sale.query.order_by(Sale.sale_date.asc())
    label = 'All Time'
    if start_str and end_str:
        try:
            start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
            query = query.filter(
                func.date(Sale.sale_date) >= start_date,
                func.date(Sale.sale_date) <= end_date
            )
            label = f'{start_str} to {end_str}'
        except ValueError:
            return jsonify({'error': 'Invalid date format.'}), 400

    sales_data = query.all()

    if fmt == 'excel':
        return _export_excel(sales_data, label)
    else:
        return _export_pdf(sales_data, label)


def _export_excel(sales_data, label):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sales Report'

    ws.merge_cells('A1:K1')
    ws['A1'] = f'Sales Report — {label}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:K2')
    ws['A2'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['A2'].alignment = Alignment(horizontal='center')

    headers = ['Sale ID', 'Date', 'Product', 'Variant', 'Qty', 'Unit Price',
               'Buying Price', 'Revenue', 'Cost', 'Profit', 'Payment Method']
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    total_revenue = total_cost = total_profit = 0
    for row_idx, s in enumerate(sales_data, 5):
        ws.cell(row=row_idx, column=1, value=s.sale_id)
        ws.cell(row=row_idx, column=2, value=s.sale_date.strftime('%Y-%m-%d %H:%M') if s.sale_date else '')
        ws.cell(row=row_idx, column=3, value=s.product.product_name if s.product else '')
        ws.cell(row=row_idx, column=4, value=s.variant._label() if s.variant else '')
        ws.cell(row=row_idx, column=5, value=s.quantity_sold)
        ws.cell(row=row_idx, column=6, value=float(s.selling_price_at_sale))
        ws.cell(row=row_idx, column=7, value=float(s.buying_price_at_sale))
        ws.cell(row=row_idx, column=8, value=float(s.total_revenue))
        ws.cell(row=row_idx, column=9, value=float(s.total_cost))
        ws.cell(row=row_idx, column=10, value=float(s.profit))
        ws.cell(row=row_idx, column=11, value=s.payment_method.method_name if s.payment_method else '')
        total_revenue += float(s.total_revenue)
        total_cost += float(s.total_cost)
        total_profit += float(s.profit)

    total_row = len(sales_data) + 5
    ws.cell(row=total_row, column=3, value='TOTALS').font = Font(bold=True)
    for col, val in [(8, total_revenue), (9, total_cost), (10, total_profit)]:
        cell = ws.cell(row=total_row, column=col, value=val)
        cell.font = Font(bold=True)

    col_widths = [8, 18, 25, 20, 6, 12, 14, 12, 12, 12, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f'sales_report_{label.replace(" ", "_").replace(":", "")}.xlsx'
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


def _export_pdf(sales_data, label):
    from reportlab.lib.pagesizes import landscape, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4),
                            leftMargin=1.5*cm, rightMargin=1.5*cm,
                            topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Title'], fontSize=16, spaceAfter=6)
    sub_style = ParagraphStyle('sub', parent=styles['Normal'], fontSize=9,
                               textColor=colors.grey, spaceAfter=12)

    story = [
        Paragraph(f'Sales Report — {label}', title_style),
        Paragraph(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")} | Total records: {len(sales_data)}', sub_style),
    ]

    total_revenue = sum(float(s.total_revenue) for s in sales_data)
    total_cost    = sum(float(s.total_cost) for s in sales_data)
    total_profit  = sum(float(s.profit) for s in sales_data)

    summary_data = [
        ['Total Revenue', 'Total Cost', 'Total Profit', 'Total Sales'],
        [f'KSh{total_revenue:,.2f}', f'KSh{total_cost:,.2f}',
         f'KSh{total_profit:,.2f}', str(len(sales_data))]
    ]
    summary_table = Table(summary_data, colWidths=[6*cm]*4)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2C3E50')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#EBF5FB')),
        ('FONTNAME', (0,1), (-1,1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.4*cm))

    headers = ['ID', 'Date & Time', 'Product', 'Variant', 'Qty',
               'Unit Price', 'Revenue', 'Cost', 'Profit', 'Payment']
    table_data = [headers]
    for s in sales_data:
        table_data.append([
            str(s.sale_id),
            s.sale_date.strftime('%Y-%m-%d %H:%M') if s.sale_date else '',
            s.product.product_name if s.product else '',
            s.variant._label() if s.variant else '',
            str(s.quantity_sold),
            f'KSh{float(s.selling_price_at_sale):,.2f}',
            f'KSh{float(s.total_revenue):,.2f}',
            f'KSh{float(s.total_cost):,.2f}',
            f'KSh{float(s.profit):,.2f}',
            s.payment_method.method_name if s.payment_method else '',
        ])

    col_widths = [1.2*cm, 3.5*cm, 4.5*cm, 3.5*cm, 1.2*cm, 3*cm, 3*cm, 3*cm, 3*cm, 3*cm]
    sales_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    sales_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#34495E')),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 8),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('ALIGN', (2,1), (2,-1), 'LEFT'),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F2F3F4')]),
        ('GRID', (0,0), (-1,-1), 0.3, colors.HexColor('#BDC3C7')),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    story.append(sales_table)

    doc.build(story)
    output.seek(0)
    filename = f'sales_report_{label.replace(" ", "_").replace(":", "")}.pdf'
    return send_file(output, mimetype='application/pdf',
                     as_attachment=True, download_name=filename)


# ─────────────────────────────────────────
# REPORTS — daily / monthly / yearly
# ─────────────────────────────────────────
@app.route('/api/reports/daily', methods=['GET'])
def daily_report():
    today = date.today()
    sales_today = Sale.query.filter(func.date(Sale.sale_date) == today).all()
    totals = db.session.query(
        func.sum(Sale.total_revenue), func.sum(Sale.total_cost),
        func.sum(Sale.profit), func.sum(Sale.quantity_sold)
    ).filter(func.date(Sale.sale_date) == today).first()

    payment_breakdown = db.session.query(
        PaymentMethod.method_name,
        func.sum(Sale.total_revenue), func.sum(Sale.profit),
        func.count(Sale.sale_id)
    ).join(Sale, PaymentMethod.payment_method_id == Sale.payment_method_id)\
     .filter(func.date(Sale.sale_date) == today)\
     .group_by(PaymentMethod.method_name).all()

    return jsonify({
        'date': today.isoformat(),
        'summary': {
            'total_revenue': float(totals[0] or 0),
            'total_cost': float(totals[1] or 0),
            'total_profit': float(totals[2] or 0),
            'total_items_sold': int(totals[3] or 0)
        },
        'sales': [s.to_dict() for s in sales_today],
        'payment_breakdown': [
            {'method': m, 'revenue': float(r), 'profit': float(p), 'transactions': int(c)}
            for m, r, p, c in payment_breakdown
        ]
    })


@app.route('/api/reports/monthly', methods=['GET'])
def monthly_report():
    year = request.args.get('year', type=int)
    month = request.args.get('month', type=int)
    if not year or not month:
        return jsonify({'error': 'year and month are required'}), 400

    totals = db.session.query(
        func.sum(Sale.total_revenue), func.sum(Sale.total_cost),
        func.sum(Sale.profit), func.sum(Sale.quantity_sold)
    ).filter(
        func.extract('year', Sale.sale_date) == year,
        func.extract('month', Sale.sale_date) == month
    ).first()

    daily_breakdown = db.session.query(
        func.date(Sale.sale_date).label('day'),
        func.sum(Sale.total_revenue), func.sum(Sale.profit), func.sum(Sale.quantity_sold)
    ).filter(
        func.extract('year', Sale.sale_date) == year,
        func.extract('month', Sale.sale_date) == month
    ).group_by(func.date(Sale.sale_date)).order_by(func.date(Sale.sale_date)).all()

    payment_breakdown = db.session.query(
        PaymentMethod.method_name,
        func.sum(Sale.total_revenue), func.sum(Sale.profit), func.count(Sale.sale_id)
    ).join(Sale, PaymentMethod.payment_method_id == Sale.payment_method_id).filter(
        func.extract('year', Sale.sale_date) == year,
        func.extract('month', Sale.sale_date) == month
    ).group_by(PaymentMethod.method_name).all()

    month_names = {1:'January',2:'February',3:'March',4:'April',5:'May',6:'June',
                   7:'July',8:'August',9:'September',10:'October',11:'November',12:'December'}

    return jsonify({
        'year': year, 'month': month,
        'summary': {
            'total_revenue': float(totals[0] or 0), 'total_cost': float(totals[1] or 0),
            'total_profit': float(totals[2] or 0), 'total_items_sold': int(totals[3] or 0)
        },
        'daily_breakdown': [
            {'date': str(d), 'revenue': float(r), 'profit': float(p), 'items_sold': int(i)}
            for d, r, p, i in daily_breakdown
        ],
        'payment_breakdown': [
            {'method': m, 'revenue': float(r), 'profit': float(p), 'transactions': int(c)}
            for m, r, p, c in payment_breakdown
        ]
    })


@app.route('/api/reports/yearly', methods=['GET'])
def yearly_report():
    year = request.args.get('year', type=int)
    if not year:
        return jsonify({'error': 'year is required'}), 400

    totals = db.session.query(
        func.sum(Sale.total_revenue), func.sum(Sale.total_cost),
        func.sum(Sale.profit), func.sum(Sale.quantity_sold)
    ).filter(func.extract('year', Sale.sale_date) == year).first()

    monthly_breakdown = db.session.query(
        func.extract('month', Sale.sale_date).label('month'),
        func.sum(Sale.total_revenue), func.sum(Sale.profit), func.sum(Sale.quantity_sold)
    ).filter(func.extract('year', Sale.sale_date) == year)\
     .group_by(func.extract('month', Sale.sale_date))\
     .order_by(func.extract('month', Sale.sale_date)).all()

    month_names = {1:'January',2:'February',3:'March',4:'April',5:'May',6:'June',
                   7:'July',8:'August',9:'September',10:'October',11:'November',12:'December'}

    return jsonify({
        'year': year,
        'summary': {
            'total_revenue': float(totals[0] or 0), 'total_cost': float(totals[1] or 0),
            'total_profit': float(totals[2] or 0), 'total_items_sold': int(totals[3] or 0)
        },
        'monthly_breakdown': [
            {'month': int(m), 'month_name': month_names[int(m)],
             'revenue': float(r), 'profit': float(p), 'items_sold': int(i)}
            for m, r, p, i in monthly_breakdown
        ]
    })


# ─────────────────────────────────────────
# PAYMENT METHODS & PAYBILLS
# ─────────────────────────────────────────
@app.route('/api/payment-methods', methods=['GET', 'POST'])
def get_payment_methods():
    if request.method == 'POST':
        data = request.get_json()
        method = PaymentMethod(method_name=data['method_name'])
        db.session.add(method)
        db.session.commit()
        return jsonify(method.to_dict()), 201
    return jsonify([m.to_dict() for m in PaymentMethod.query.all()])


@app.route('/api/paybills', methods=['GET', 'POST'])
def paybills():
    if request.method == 'POST':
        data = request.get_json()
        new_paybill = Paybill(
            paybill_name=data['paybill_name'],
            paybill_number=data['paybill_number'],
            payment_method_id=data['payment_method_id']
        )
        db.session.add(new_paybill)
        db.session.commit()
        return jsonify(new_paybill.to_dict()), 201
    return jsonify([p.to_dict() for p in Paybill.query.all()])


@app.route('/api/payment-methods/<int:method_id>', methods=['DELETE'])
def delete_payment_method(method_id):
    method = PaymentMethod.query.get_or_404(method_id)
    if Sale.query.filter_by(payment_method_id=method_id).first():
        return jsonify({'error': f'Cannot delete "{method.method_name}" — it has sales records attached.'}), 409
    Paybill.query.filter_by(payment_method_id=method_id).delete()
    db.session.delete(method)
    db.session.commit()
    return jsonify({'message': f'"{method.method_name}" deleted.'}), 200


@app.route('/api/paybills/<int:paybill_id>', methods=['DELETE'])
def delete_paybill(paybill_id):
    paybill = Paybill.query.get_or_404(paybill_id)
    db.session.delete(paybill)
    db.session.commit()
    return jsonify({'message': f'"{paybill.paybill_name}" removed.'}), 200


# ─────────────────────────────────────────
# PAGE ROUTES
# ─────────────────────────────────────────
@app.route('/')
def home():
    return redirect(url_for('inventory_page'))

@app.route('/inventory')
def inventory_page():
    return render_template('inventory.html')

@app.route('/products/new')
def new_product_page():
    return render_template('product_form.html')

@app.route('/products/<int:product_id>/edit')
def edit_product_page(product_id):
    return render_template('product_form.html', product_id=product_id)

@app.route('/sales/new')
def record_sale_page():
    return render_template('record_sale.html')

@app.route('/reports/daily')
def daily_report_page():
    return render_template('daily_report.html')

@app.route('/reports')
def reports_page():
    return render_template('reports.html')

@app.route('/settings')
def settings_page():
    return render_template('settings.html')


if __name__ == '__main__':
    app.run(debug=True)